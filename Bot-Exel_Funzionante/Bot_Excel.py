from openpyxl import *
import telegram
import json
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters
from tok import *
filename ="DataBase.xlsx" #nome del file exel da elaborare
sheetT = "Database" #nome del foglio da elaborare
sheetTNKW = "Nk" #nome del foglio per le parole non ancora definite
TOKEN = "" #token di BotFather
dir = load_workbook(filename)
sheet = dir[sheetT]

#Funzione che aggiunge la parola non presente nel database al file.
def addNKW(s):
    try:
        #sh = dir.get_sheet_by_name(sheetTNKW)
        sh = dir[sheetTNKW]
    except:
        #print("Non posso lavorare sul file perchè è già aperto dal sistema operativo!")
        exit
    r = 1
    c = 1
    rs = 0
    cmp = True
    while(sh.cell(r, c).value != None):
        if(sh.cell(r, c).value == s): #controllo per vedere se la parola da inserire in lista è già presente, nel caso, non la aggiungo
            cmp = False
            rs = r
            break
        r += 1
    if cmp:
        sh.cell(r, c, s) #aggiungo al file la parola non presente nel database
        sh.cell(r, (c + 1), 1)
        dir.save(filename)# salvo il file
    else:
        volte = int(sh.cell(rs, (c + 1)).value)
        volte += 1
        sh.cell(rs, (c + 1), volte)
        dir.save(filename)
        
def getToken():
    t = tok()
    return t.getToken()
#def readFile(v, rows, coloumns): 
    #v[0] = sheet.cell(rows, coloumns).value

#Funzione che cerca la parola nel foglio excel, se la trova, stampa in chat la sua definzione, altrimenti, la aggiunge nel foglio delle parole non trovate
# e stampa in chat un messaggio per avvisare che la parola cercata non si trova nel database.
def findWord(stri, up, cont):
    #A differenza delle classiche matrici, i fogli excel partono da 1,1
    r = 2
    c = 1
    v = None
    defi = None
    while(sheet.cell(r, c).value != None):
        if(sheet.cell(r, c).value == stri):
            v = sheet.cell(r, c).value
            defi = sheet.cell(r, (c + 1)).value
            break
        else:
            r += 1
            v = None
    if(v == None):
        up.message.reply_text("\nLa parola non è ancora presente nell'archivio!\nIl nostro team la aggiungerà il prima possibile!")
        addNKW(stri)
    else:
        up.message.reply_text("Parola: "+ str(v) +"\nDefinizione: " + str(defi))    

#Funzione che serve ad inizializzare i fogli excel
def initRead(f, sh):
    dir = load_workbook(f)
    sheet = dir[sh]

#Funzione che viene lanciata al comando /define parola 
def defineFunction(up, cont):
    initRead(filename, sheetT)# inizializza il foglio excel
    string  = inputBot(up, cont)  #ottiene la parola da ricercare
    if(len(string) == 0 or string == "-1"): #se la parola è vuota o si è verificato qualche errore nell'input stampa il seguente errore
        up.message.reply_text("Non hai inserito una parola da ricercare!\nSintassi: /def parola_da_definire")
    else: #altrimenti cerca la parola nel database 
        findWord(string, up, cont)#cerca la parola nel file excel

#Funzione che, dato il comando, estrapola e fornisce la parola/le parole da cercare nel database. 
def inputBot(up, cont):
    buf = up.message.text #ottiene tutta la stringa del comando
    dele = "/def" 
    if(buf == dele):
        return "-1"
    la = buf.split(" ")
    lb = dele.split(" ")
    lc = [x for x in la if x not in lb]#ciclo per gestire le parole con lo spazio ("San Valentino")
    stri = " ".join(lc)# Prende tutte le parole presenti nella lista, le unisce in una sola stringa e le separa con il carattere " " ('32[10]', '0x20')
    stri = stri.casefold()#fa diventare tutta la stringa minuscola
    stri = stri.capitalize()#fa diventare la prima lettera maiuscola
    up.message.reply_text("Parola da ricercare: " + stri)
    return stri

#Funzione lanciata al comando /start
def start(up, cont):
    info = up.message.from_user
    nome = info['first_name']
    up.message.reply_text("Ciao " + str(nome) +  ", per definire una parola digita: \n/def parola_da_definire (ES: /def virus)")


def main():
    print("In esecuzione! ")
    TOKEN = getToken()
    upd= Updater(TOKEN, use_context=True)
    disp=upd.dispatcher
    disp.add_handler(CommandHandler("start", start))
    disp.add_handler(CommandHandler("def", defineFunction))
    upd.start_polling()#Chiede a telegram se ci sono nuovi messaggi
    upd.idle() #permette al bot di smettere la sua esecuzione tramite shortcut da tastiera CRTL + C (interrupt)

if __name__=='__main__':
   main() #Lancia la funzione main