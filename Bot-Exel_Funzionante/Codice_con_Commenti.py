#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

""""Importazione delle librerie e dei moduli utili.
La libreria telegram serve per gestire i collegamenti con telegram per esempio e le varie fuznnzioni.
'openpyxl' è la libreria che ci permette di ineragire con i fogli di calcolo di excel che in questo caso sono il nostro database temporaneo."""
from openpyxl import *
import telegram
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Dichiaro una variabile filename che contiene il nome del file excel che ci fa da database
filename = "DataBase.xlsx" #nome del file exel da elaborare

#Dichiaro una variabile sheetT che contiene il nome del foglio di lavoro(del file excel) che contiene le parole salvate
sheetT = "Database" #nome del foglio da elaborare

#Dichiaro una variabile sheetTNKW che contiene il nome del foglio di lavoro(del file excel) che contiene ancora NON salvate
sheetTNKW = "Nk" #nome del foglio per le parole non ancora definite

"""Questo è il token che ci viene dato da botfather. Botfather è un canale telegram svilupato da loro stessi, che ci aiuta nella creazione
di un nostro bot. Il token invece è un codice che identifica il nostro bot."""
TOKEN = "1657410885:AAHS56NZi2fypJ3Ia2VyDZvb7_JrhlqiV4I" 

"""La funzione dir è una funzione che mi permette di inizializzare i fogli di excel, infatti è presa dalla libreria openpyxl
load_workbook è sempre un comando ripreso dalla libreria openpyxl, ma serve ad aprire un file excel. Quando lo apre crea una lista di fogli e poi con il comando dir vado ad accedere al singolo foglio"""
dir = load_workbook(filename)
sheet = dir[sheetT]

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

"""Funzione che serve per aggiungere al foglio di excel contenente le parole senza definizione, la parola cercata. 
Effettuato il passaggio di parametro di s, che sarebbbe la parola ricercata e non trovata nel database."""
def addNKW(s):

    """Handler try-except per gestire le eccezioni. Quello indentato a try sono le azioni che vengono eseguite normalmente, mentre quelle indentate a except
    sono le azioni che esegue in caso di eccezioni"""
    try:
        #sh = dir.get_sheet_by_name(sheetTNKW)
        sh = dir[sheetTNKW]
    except:
        #L'eccezione in questo caso è quella che il file sia aperto dall'utente.  Se si verifica esce dalla funzione, non aggiunge la parola al database, ed il bot ritorna in ascolto
        #print("Non posso lavorare sul file perchè è già aperto dal sistema operativo!")
        exit

    #Dichiaro e inizializzazio le variabili r e c. r identifica le righe, mentre c le colonne. Entrambe le inizializzo a 1
    r = 1
    c = 1

    #Dichiaro una variabile di tipo booleana che si chiama cmp e la inizializzo a True
    cmp = True

    #Creo un ciclo while con condizione che la cella in cui sono sia diversa da None, che non sia vuota
    while(sh.cell(r, c).value != None):
        if(sh.cell(r, c).value == s): #controllo per vedere se la parola da inserire in lista è già presente, nel caso, non la aggiungo
            cmp = False
            break
        r += 1
    
    #Questo lo fa se la variabile cmp è impostata su False
    if cmp:

        #aggiungo al file la parola non presente nel database
        sh.cell(r, c, s) 

        #Salvo il file excel
        dir.save(filename)
        
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#def readFile(v, rows, coloumns): 
    #v[0] = sheet.cell(rows, coloumns).value

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

"""Funzione che cerca la parola nel foglio excel, se la trova, stampa in chat la sua definzione, altrimenti, la aggiunge nel foglio delle parole non trovate
e stampa in chat un messaggio per avvisare che la parola cercata non si trova nel database."""
def findWord(stri, up, cont):
    
    """A differenza delle classiche matrici, i fogli excel partono da 1,1
    Quindi dichiaro due variabili che sono r e c. La prima mi identifica le righe e la seconda le colonne.
    Le righe le inizializzo a due perchè la prima, se guardiamo, è di "descrizione" del foglio di lavoro. Le colonne invece la inizializzo ad uno perchè è quella in cui stanno le parole"""
    r = 2
    c = 1

    #Dichiaro una variabile di nome v e la inizializzo a None
    v = None

    #Dichiaro una variabile di nome defi e la inizializzo a None
    defi = None

    #Creo un ciclo while con condizione che la cella in cui sono sia diversa da None, ossia che non è vuota
    while(sheet.cell(r, c).value != None):

        #Controllo se il valore della cella in cui sono è uguale alla stringa contenuta nella variabile stri
        if(sheet.cell(r, c).value == stri):
            #Se le stringhe corrispondono metto nella variabile v il contenuto della cella in cui sono
            v = sheet.cell(r, c).value

            #Dopodichè nella variabile defi (che sta per definizione) ci vado a mettere ciò che è contenuto nella cella della stessa riga, ma della colonna successiva
            defi = sheet.cell(r, (c + 1)).value

            #Infine viene terminata l'esecuzione del programma 
            break
        
        #Altrimenti se le stringhe non corrispondono fa le due azioni che seguono
        else:
            #Vado ad incrementare le righe di uno
            r += 1

            #Riassegno un valore a v che è None
            v = None

    #Costrutto if che mi controlla (ossia ha come condizione) se la parola ricercata è presente o meno (ossia v == None)
    if(v == None):
        up.message.reply_text("\nLa parola non è ancora presente nell'archivio!\nIl nostro team la aggiungerà il prima possibile!")
        
        #Viene richiamata la funzione addKNW vista precedentemente, con passaggio di parametro della variabile stri, che sarebbe la parola ricercata.
        addNKW(stri)
    else:
        """Se la variabile v non è impostata su None (ossia la parola è stata trovata) viene fatta l'azione che segue, ossia la stampa della definizione della parola"""
        up.message.reply_text("Parola: "+ str(v) +"\nDefinizione: " + str(defi))    

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Funzione che serve ad inizializzare i fogli excel
def initRead(f, sh):
    dir = load_workbook(f)
    sheet = dir[sh]

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Funzione che viene lanciata al comando /define parola_da_definire
def defineFunction(up, cont):
    #Inizializzo il foglio di excel, andando a specificare tra i parametri il file excel in questione e lo specifico foglio di lavoro
    initRead(filename, sheetT) # inizializza il foglio excel

    #Chiedo da input la parola da ricercare
    string  = inputBot(up, cont)  #ottiene la parola da ricercare

    """Costrutto if per verificare che ciò che l'utente ha scritto vada bene
    Se la parola (la variabile string) è vuota o si è verificato qualche errore nell'input stampa il seguente errore"""
    if(len(string) == 0 or string == "-1"): 
        #Messaggio mandato come risposta dal bot
        up.message.reply_text("Non hai inserito una parola da ricercare!\nSintassi: /def parola_da_definire")
        #Altrimenti se non ci sono errori in ciò che l'utente ha scritto, si va a ricercare la parola
    else: 
        #Viene richiamata la funzione findWord che cerca la parola richiesta nel file excel. Vengono passati tre parametri.
        findWord(string, up, cont)

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Funzione che, dato il comando, estrapola e fornisce la parola/le parole da cercare nel database. 
def inputBot(up, cont):

    #Dentro la variabile buf mi va l'intero comando inserito dall'utente (per esempio: /def San Valentino)
    buf = up.message.text #ottiene tutta la stringa del comando

    #Dichiaro una variabile dele che contiene la stringa /def
    dele = "/def" 

    #Se le stringhe contenute in buf e dele sono uguali allora c'è stato un errore
    if(buf == dele):
        return "-1"

    #Divido la stringa buf in tante sottostringhe, quest'ultime vengono create dividendo la scritta principale tutte le volte che si incontra il 
    #carattere passato come parametro alla funzione (in questo caso lo spazio).
   
    la = buf.split(" ")

    #Eseguo la stessa operazione del rigo precedente però sulla variabile lb. 
    # In questo caso la parola è una sola, quindi il cambiamento è pressochè inesistente, ma è necessario per l'istruzione successiva: 
    # Nel caso si ometta questa istruzione, si avranno problemi relativi al casting della variabile, ovvero al tipo di variabile che sto utilizzando.
    
    lb = dele.split(" ")

    """Creo una variabile lc e dico che il suo contenuto è il risultato delle operazion svolte all'interno delle parentesi quadre.
    Ciò che andrà dentro la variabile lc è ogni elemento (x) della variabile la che NON è uguale al contenuto della variabile lb. 
    Continuando a seguire l'esempio fatto fino ad ora, dentro la variabile lc avremo le parole San e Valentino, poichè sono le parole che sono presenti nella variabile la,
    ma non nella variabile lb. 
    La parola /def non verrà caricata dentro lc poichè è una parola comune alle variabili la e lb
    Brevemente elimino il comando digitato dall'utente (/def) e gestisco anche il caso in cui le stringhe sono composte da più parole separate da uno spazio (San Valentino)
    """
    lc = [x for x in la if x not in lb] #ciclo per gestire le parole con lo spazio ("San Valentino")

    """Creo una variabile di nome stri e ci vado a mettere il contenuto della variabile lc: 
        Grazie alla funzione join vado a concatenare le singole parole contenute nella lista e le separo con uno spazio."""
    stri = " ".join(lc) # Prende tutte le parole presenti nella lista, le unisce in una sola stringa e le separa con il carattere ('32[10]', '0x20')
    
    #Fa diventare tutta la stringa contenuta all'interno della variabile stri minuscola
    stri = stri.casefold()
    
    #Fa diventare la prima lettera della stringa contenuta all'interno della variabile stri maiuscola
    stri = stri.capitalize()

    #Stampa come risposta del bot la parola da ricercare
    up.message.reply_text("Parola da ricercare: " + stri)

    #Viene restituito stri
    return stri

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Funzione lanciata quando viene digitato il comando: /start
def start(up, cont):
    
    info = up.message.from_user

    #Dentro alla variabile nome ci va a finire il nome dell'utente che sta usando il bot
    nome = info['first_name']

    #Messaggio mandato come risposta dal bot
    up.message.reply_text("Ciao " + str(nome) +  ", per definire una parola digita: \n/def parola_da_definire (ES: /def armadio)")

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Funzione principale richiamata quando lancio il programma e il bot
def main():
    #Stampa sul terminale della stringa tra virgolette tra parentesi
    print("In esecuzione! ")

    """Creazione di un Updater a cui viene passato il token del bot telegram. Un Updater come dice la parola stessa serve a ricevere gli aggiornamenti da Telegram.
    use_context=True è opzionale e se impostato su true (come nel nostro caso) utilizza l'API di callback basata sul contesto.
    La variabile upd adesso è considerato come un oggetto, da cui di conseguenza posso richiamare vari metodi"""
    upd= Updater(TOKEN, use_context=True)

    #Dall'oggetto upd richiamo il metodo dispatcher che invia tutti i tipi di aggiornamenti ai suoi gestori registrati
    disp=upd.dispatcher

    """Il comando add_handler aggiunge dei comandi al nostro bot di telegram.
    Il primo parametro è il comando che bisogna digitare per fare una determinata azione, mentre il seondo è la funzione che viene richiamata una volta digitato il comando"""
    disp.add_handler(CommandHandler("start", start))
    disp.add_handler(CommandHandler("def", defineFunction))

    #Richiamo del metodo start_polling, che avvia una procedura che ci fornisce eventuali aggiornamenti sulla chat Telegram
    upd.start_polling()#Chiede a telegram se ci sono nuovi messaggi

    """Richiamo del metodo idle() che termina l'esecuzione del bot, di default, con la combinazione si tasti Ctrl+C.
    Volendo con il parametro stop_signals= <> , se tra le parentesi angolari mettiamo una combinazione di tasti, quella diventerà la nuova combinazione per stopparlo"""
    upd.idle() #permette al bot di smettere la sua esecuzione tramite shortcut da tastiera CRTL + C (interrupt)

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#Dichiarazione del programma principale e richiamo della funzione main()
if __name__=='__main__':
   main() #Lancia la funzione main

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------