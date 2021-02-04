from openpyxl import *
import telegram
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters
filename = "DataBase.xlsx" #nome del file exel da elaborare
sheetT = "Database" #nome del foglio da elaborare
sheetTNKW = "Nk" #nome del foglio per le parole non ancora definite
TOKEN = "1524897969:AAFea-xo8G-xQvuoBhGUWXjQcdNAg7kQQpw"
dir = load_workbook(filename)
sheet = dir[sheetT]

def addNKW(s):
    try:
        #sh = dir.get_sheet_by_name(sheetTNKW)
        sh = dir[sheetTNKW]
    except:
        #print("Non posso lavorare sul file perchè è già aperto dal sistema operativo!")
        exit
    r = 1
    c = 1
    cmp = True
    while(sh.cell(r, c).value != None):
        if(sh.cell(r, c).value == s):
            cmp = False
            break
        r += 1
    r = 1
    if cmp:
        while(sh.cell(r, c).value != None):
            r += 1
        sh.cell(r, c, s)
        dir.save(filename)
        
def readFile(v, rows, coloumns): 
    v[0] = sheet.cell(rows, coloumns).value

def findWord(stri, up, cont):
    r = 2
    c = 1
    v = None
    defi = None
    while(sheet.cell(r, c).value != None):
        if(sheet.cell(r, c).value == stri):
            v = sheet.cell(r, c).value
            defi = sheet.cell(r, (c + 1)).value
            break
        else:
            r += 1
            v = None
    if(v == None):
        up.message.reply_text("\nLa parola non è ancora presente nell'archivio!\nIl nostro team la aggiungerà il prima possibile!")
        addNKW(stri)
    else:
        up.message.reply_text("Parola: "+ str(v) +"\nDefinizione: " + str(defi))    

def initRead(f, sh):
    dir = load_workbook(f)
    sheet = dir[sh]

def defineFunction(up, cont):
    initRead(filename, sheetT)
    string  = inputBot(up, cont)
    if(len(string) == 0 or string == "-1"):
        up.message.reply_text("Non hai inserito una parola da ricercare!\nSintassi: /def parola_da_definire")
    else:
        findWord(string, up, cont)

def inputBot(up, cont):
    buf = up.message.text
    dele = "/def"
    if(buf == dele):
        return "-1"
    la = buf.split(" ")
    lb = dele.split(" ")
    lc = [x for x in la if x not in lb]
    stri = " ".join(lc)
    stri = stri.casefold()
    stri = stri.capitalize()
    up.message.reply_text("Parola da ricercare: " + stri)
    return stri

def start(up, cont):
    info = up.message.from_user
    nome = info['first_name']
    up.message.reply_text("Ciao " + str(nome) +  ", per definire una parola digita: \n/def parola_da_definire (ES: /def armadio)")


def main():
    upd= Updater(TOKEN, use_context=True)
    disp=upd.dispatcher
    disp.add_handler(CommandHandler("start", start))
    disp.add_handler(CommandHandler("def", defineFunction))
    upd.start_polling()
    upd.idle()

if __name__=='__main__':
   main()